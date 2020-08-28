a = []
b = [1, 2, 3]
c = ['Life', 'is', 'too', 'short']
d = [1, 2, 'life', 'is']
e = [1, 2, ['List', 'is']]

print(b[2])
print(d[2][1])
print(e[2][1])
print(b[-1])  # 마지막 리스트 요소
print(c[1:3])  # 마지막 인덱스 -1
print(c[:3])
print(c[3:])

a.append("추가")
a.append("두번째추가")
print(a)

del a[0]
print(a)

##########################################

myList = "This is a book That is a pencil".split()
print(myList)
i = myList.index('book')
n = myList.count('is')
print("리스트 요소에서 book 인덱스: %d 리스트 요소에서의 is 갯수: %s" % (i, n))

#############################################
# 요소 병합, 반복
a = [1, 2]
b = [3, 4, 5]
c = a + b
print(c)

d = a * 3
print(d)

# reverse 리스트 뒤집기(거꾸로)
c.reverse()
print(c)
# 정렬
c.sort()
print(c)

#############################################
# dictionary
scores = {'선영': 90, ',햔수': 50, '행복하쟈': 40}

for key in scores:
    var = scores[key]
    print("%s : %d" % (key, var))

for a in scores:
    var = scores[a]
    print("%s : %d" % (a, var))

# tuple list로부터 dict 생성
people = [('선영', 90), (',햔수', 50), ('행복하쟈', 40)]
print(type(people), people)
mydict = dict(people)
print(type(mydict))
age = mydict['선영']
print(mydict)
print(age)
######################
import random

ran1 = random.random()
print(ran1)

print(random.randrange(1, 7))
abc = ['a', 'b', 'c', 'd', 'e']
random.shuffle(abc)
print(abc)
print(random.choice(abc))

######################
# 입력받은 수의 1~10 사이의 임의수가 같을 경오 '통과' 출력 아닐경우 될때까지 입력 받으시오
"""
import random

ran1 = random.randrange(1, 10)
# print(ran1)
num = input("숫자를 입력하세요: ")

print(num)
while 1:
    if ran1 == int(num):
        print("통과")
        break;
    else:
        num = input("숫자를 입력하세요: ")
"""

######################
"""import numpy as np

ran1 = np.random.randint(1, 10)
print(ran1)
num = input("1 ~ 10숫자를 입력하세요(랜덤 값과 같아야 출력됨): ")

while 1:
    if ran1 == int(num):
        print("통과")
        break
    else:
        if ran1 > int(num):
            print("더 큰 수")
        else:
            print("더 작은 수")
        num = input("1 ~ 10숫자를 입력하세요(랜덤 값과 같아야 출력됨): ")"""

######################
"""from openpyxl import load_workbook

# data_only=True로 해줘야 수식이 아닌 값으로 받아온다
load_wb = load_workbook("emp.xlsx", data_only=True)

# 시트 이름으로 불러오기
load_ws = load_wb['sheet1']
print(load_ws.cell(row=1, column=1).value)
print(load_ws.cell(row=1, column=2).value)
print(load_ws.cell(row=1, column=3).value)

# 셀 좌표로 값 출력
print("셀 좌표로 값 출력: ", load_ws.cell(2, 2).value)
print("셀 좌표로 값 출력: ", load_ws.cell(2, 3).value)
all_values = []
for row in load_ws.rows:
    row_value = []
    for cell in row:
        row_value.append(cell.value)
    all_values.append(row_value)
print(all_values)

list1 = [10, 100, 49]
print(max(list1))

for i, row in enumerate(load_ws.rows):
    if i == 0:
        print(row[0].value)
        print(row[1].value)
        print(row[2].value)"""

